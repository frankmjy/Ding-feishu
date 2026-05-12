from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import re
import shutil
import subprocess
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, unquote, urlparse

import requests
from openpyxl import load_workbook

try:
    from dotenv import load_dotenv
except ImportError:  # pragma: no cover - used only when dependencies are missing.
    load_dotenv = None


EXCEL_MIME_HINTS = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
    "application/octet-stream",
)

READ_ONLY_FIELD_TYPES = {
    19,  # formula
    20,  # lookup / duplex-related fields in older API responses
    21,
    22,
    1001,  # created time
    1002,  # modified time
    1003,  # created by
    1004,  # modified by
}


@dataclass(frozen=True)
class Settings:
    dingtalk_doc_url: str
    dingtalk_doc_title: str
    feishu_bitable_url: str
    feishu_app_id: str
    feishu_app_secret: str
    feishu_bitable_app_token: str | None
    local_excel_path: Path
    sync_mode: str
    field_mapping_file: Path | None
    sheet_name: str | None
    header_row: int
    create_missing_fields: bool
    dingtalk_download_mode: list[str]
    dingtalk_browser_profile: Path
    dingtalk_browser_timeout_sec: int
    dingtalk_auto_click_export: bool
    dingtalk_auto_login: bool
    dingtalk_export_kind: str
    dingtalk_import_browser_cookies: bool
    dingtalk_browser_cookie_source: str
    dingtalk_login_hint: str
    feishu_api_base: str
    feishu_notify_enabled: bool
    feishu_notify_webhook_url: str | None
    feishu_notify_webhook_secret: str | None
    feishu_notify_app_id: str | None
    feishu_notify_app_secret: str | None
    feishu_notify_chat_id: str | None
    feishu_notify_open_id: str | None
    feishu_member_lookup_chat_id: str | None


@dataclass(frozen=True)
class SyncResult:
    title: str
    excel_rows: int
    created: int = 0
    updated: int = 0
    skipped: int = 0
    mode: str = ""
    target_url: str = ""


class SyncError(RuntimeError):
    pass


class FeishuAPIError(SyncError):
    pass


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Download a DingTalk document as Excel and sync it to Feishu Bitable."
    )
    parser.add_argument("--env", default=".env", help="Path to the .env file.")
    parser.add_argument("--excel", help="Override LOCAL_EXCEL_PATH.")
    parser.add_argument("--sheet", help="Override SHEET_NAME.")
    parser.add_argument("--header-row", type=int, help="Override HEADER_ROW.")
    parser.add_argument(
        "--mode",
        choices=("insert_missing", "upsert", "replace", "append"),
        help="Override SYNC_MODE.",
    )
    parser.add_argument("--skip-download", action="store_true", help="Use the local Excel file only.")
    parser.add_argument("--download-only", action="store_true", help="Download Excel and stop.")
    parser.add_argument("--dry-run", action="store_true", help="Parse Excel but do not call Feishu APIs.")
    parser.add_argument(
        "--create-missing-fields",
        choices=("true", "false"),
        help="Override CREATE_MISSING_FIELDS.",
    )
    return parser.parse_args()


def load_env_file(path: Path, override: bool = True) -> None:
    if load_dotenv is not None:
        # Project-local .env should win over ambient machine variables. This
        # avoids accidentally calling Feishu with another app configured globally.
        load_dotenv(path, override=override)
        return

    if not path.exists():
        return

    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        if override or key not in os.environ:
            os.environ[key] = value.strip().strip('"').strip("'")


def env_bool(name: str, default: bool) -> bool:
    value = os.getenv(name)
    if value is None or value == "":
        return default
    return value.strip().lower() in {"1", "true", "yes", "y", "on"}


def env_int(name: str, default: int) -> int:
    value = os.getenv(name)
    if value is None or value == "":
        return default
    try:
        return int(value)
    except ValueError as exc:
        raise SyncError(f"{name} must be an integer, got {value!r}") from exc


def require_env(name: str) -> str:
    value = os.getenv(name, "").strip()
    if not value:
        raise SyncError(f"Missing required setting: {name}")
    return value


def load_settings(args: argparse.Namespace) -> Settings:
    env_path = Path(args.env)
    base_env_path = Path(".env")
    if env_path.resolve() != base_env_path.resolve() and base_env_path.exists():
        load_env_file(base_env_path, override=False)
    load_env_file(env_path, override=True)

    local_excel = Path(args.excel or os.getenv("LOCAL_EXCEL_PATH", "downloads/dingtalk_export.xlsx"))
    sheet_name = args.sheet if args.sheet is not None else os.getenv("SHEET_NAME", "").strip() or None
    header_row = args.header_row if args.header_row is not None else env_int("HEADER_ROW", 1)
    sync_mode = args.mode or os.getenv("SYNC_MODE", "replace").strip().lower()
    if sync_mode not in {"insert_missing", "upsert", "replace", "append"}:
        raise SyncError("SYNC_MODE must be one of: 'insert_missing', 'upsert', 'replace', or 'append'.")

    create_missing = env_bool("CREATE_MISSING_FIELDS", True)
    if args.create_missing_fields is not None:
        create_missing = args.create_missing_fields == "true"

    download_mode = [
        item.strip().lower()
        for item in os.getenv("DINGTALK_DOWNLOAD_MODE", "direct,browser").split(",")
        if item.strip()
    ]
    invalid_modes = set(download_mode) - {"direct", "browser"}
    if invalid_modes:
        raise SyncError(f"Unsupported DINGTALK_DOWNLOAD_MODE value(s): {', '.join(sorted(invalid_modes))}")

    export_kind = os.getenv("DINGTALK_EXPORT_KIND", "auto").strip().lower()
    if export_kind not in {"auto", "document", "bitable"}:
        raise SyncError("DINGTALK_EXPORT_KIND must be one of: auto, document, or bitable.")

    notify_webhook_url = os.getenv("FEISHU_NOTIFY_WEBHOOK_URL", "").strip() or None
    notify_app_id = os.getenv("FEISHU_NOTIFY_APP_ID", "").strip() or None
    notify_app_secret = os.getenv("FEISHU_NOTIFY_APP_SECRET", "").strip() or None
    notify_chat_id = os.getenv("FEISHU_NOTIFY_CHAT_ID", "").strip() or None
    notify_open_id = os.getenv("FEISHU_NOTIFY_OPEN_ID", "").strip() or None
    notify_enabled = env_bool(
        "FEISHU_NOTIFY_ENABLED",
        bool(notify_webhook_url or notify_chat_id or notify_open_id),
    )

    return Settings(
        dingtalk_doc_url=require_env("DINGTALK_DOC_URL"),
        dingtalk_doc_title=os.getenv("DINGTALK_DOC_TITLE", "").strip(),
        feishu_bitable_url=require_env("FEISHU_BITABLE_URL"),
        feishu_app_id=require_env("FEISHU_APP_ID"),
        feishu_app_secret=require_env("FEISHU_APP_SECRET"),
        feishu_bitable_app_token=os.getenv("FEISHU_BITABLE_APP_TOKEN", "").strip() or None,
        local_excel_path=local_excel,
        sync_mode=sync_mode,
        field_mapping_file=Path(os.getenv("FIELD_MAPPING_FILE", "")).resolve()
        if os.getenv("FIELD_MAPPING_FILE", "").strip()
        else None,
        sheet_name=sheet_name,
        header_row=header_row,
        create_missing_fields=create_missing,
        dingtalk_download_mode=download_mode,
        dingtalk_browser_profile=Path(os.getenv("DINGTALK_BROWSER_PROFILE", ".browser/dingtalk")),
        dingtalk_browser_timeout_sec=env_int("DINGTALK_BROWSER_TIMEOUT_SEC", 300),
        dingtalk_auto_click_export=env_bool("DINGTALK_AUTO_CLICK_EXPORT", False),
        dingtalk_auto_login=env_bool("DINGTALK_AUTO_LOGIN", True),
        dingtalk_export_kind=export_kind,
        dingtalk_import_browser_cookies=env_bool("DINGTALK_IMPORT_BROWSER_COOKIES", True),
        dingtalk_browser_cookie_source=os.getenv("DINGTALK_BROWSER_COOKIE_SOURCE", "auto").strip().lower(),
        dingtalk_login_hint=os.getenv("DINGTALK_LOGIN_HINT", "").strip(),
        feishu_api_base=os.getenv("FEISHU_API_BASE", "https://open.feishu.cn/open-apis").rstrip("/"),
        feishu_notify_enabled=notify_enabled,
        feishu_notify_webhook_url=notify_webhook_url,
        feishu_notify_webhook_secret=os.getenv("FEISHU_NOTIFY_WEBHOOK_SECRET", "").strip() or None,
        feishu_notify_app_id=notify_app_id,
        feishu_notify_app_secret=notify_app_secret,
        feishu_notify_chat_id=notify_chat_id,
        feishu_notify_open_id=notify_open_id,
        feishu_member_lookup_chat_id=os.getenv("FEISHU_MEMBER_LOOKUP_CHAT_ID", "").strip() or None,
    )


def ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def ps_single_quote(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def cleanup_automation_browser_processes(profile_dir: Path) -> None:
    """Close only Edge/Chrome processes that use this automation profile."""
    if os.name != "nt":
        return
    profile_marker = str(profile_dir.resolve())
    script = (
        f"$marker = {ps_single_quote(profile_marker)}; "
        "$names = @('msedge.exe','chrome.exe','msedgewebview2.exe'); "
        "Get-CimInstance Win32_Process | "
        "Where-Object { $_.CommandLine -and $_.CommandLine.Contains($marker) -and $names -contains $_.Name } | "
        "ForEach-Object { Stop-Process -Id $_.ProcessId -Force -ErrorAction SilentlyContinue }"
    )
    subprocess.run(
        ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", script],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        check=False,
    )


def mark_browser_profile_clean(profile_dir: Path) -> None:
    """Tell Chromium not to show crash/session restore UI for this profile."""
    for preferences_path in [profile_dir / "Preferences", profile_dir / "Default" / "Preferences"]:
        if not preferences_path.exists():
            continue
        try:
            data = json.loads(preferences_path.read_text(encoding="utf-8"))
            profile = data.setdefault("profile", {})
            profile["exit_type"] = "Normal"
            profile["exited_cleanly"] = True
            session = data.setdefault("session", {})
            session["restore_on_startup"] = 5
            preferences_path.write_text(json.dumps(data, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
        except Exception:
            continue


def clear_browser_session_restore(profile_dir: Path) -> None:
    """Prevent Edge/Chrome from reopening stale DingTalk preview tabs."""
    candidates = [
        profile_dir / "Current Session",
        profile_dir / "Current Tabs",
        profile_dir / "Last Session",
        profile_dir / "Last Tabs",
    ]
    for session_dir in [profile_dir / "Sessions", profile_dir / "Default" / "Sessions"]:
        if session_dir.exists():
            candidates.extend(item for item in session_dir.iterdir() if item.is_file())

    removed = 0
    for candidate in candidates:
        try:
            if candidate.exists() and candidate.is_file():
                candidate.unlink()
                removed += 1
        except OSError:
            continue
    if removed:
        print(f"[download] Cleared {removed} stale browser session restore file(s).")


def looks_like_excel(content: bytes, content_type: str | None) -> bool:
    if content.startswith(b"PK\x03\x04"):
        return True
    if content_type:
        lowered = content_type.lower()
        return any(hint in lowered for hint in EXCEL_MIME_HINTS)
    return False


def try_direct_download(url: str, output_path: Path) -> Path | None:
    print("[download] Trying direct HTTP download...")
    response = requests.get(
        url,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124 Safari/537.36"
            )
        },
        timeout=45,
        allow_redirects=True,
    )
    response.raise_for_status()
    content_type = response.headers.get("Content-Type")
    if not looks_like_excel(response.content, content_type):
        print("[download] Direct response is not an Excel file; browser login/export is required.")
        return None

    ensure_parent(output_path)
    output_path.write_bytes(response.content)
    print(f"[download] Saved direct export: {output_path}")
    return output_path


def candidate_download_path(output_path: Path) -> Path:
    suffix = output_path.suffix or ".xlsx"
    stamp = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    return output_path.with_name(f"{output_path.stem}.candidate.{stamp}{suffix}")


def visible_count(page: Any, selector: str) -> int:
    try:
        return page.locator(selector).count()
    except Exception:
        return 0


def click_first_visible(page: Any, selectors: list[str], timeout_ms: int = 2500) -> bool:
    for selector in selectors:
        try:
            locator = page.locator(selector)
            count = locator.count()
            for index in range(min(count, 12)):
                item = locator.nth(index)
                if item.is_visible(timeout=500):
                    item.click(timeout=timeout_ms)
                    return True
        except Exception:
            continue
    return False


def click_export_menu_item(page: Any) -> bool:
    return click_first_visible(
        page,
        [
            "text=/导出为\\s*Excel/i",
            "text=/下载为\\s*Excel/i",
            "text=/Excel/i",
            "text=/XLSX/i",
            "text=/导出/i",
            "text=/下载/i",
            "[aria-label*='下载']",
            "[title*='下载']",
            "[data-testid*='download']",
            "[role='menuitem']:has-text('导出')",
            "[role='menuitem']:has-text('下载')",
            "[role='option']:has-text('导出')",
            "[role='option']:has-text('下载')",
        ],
    )


def hover_first_visible(page: Any, selectors: list[str], timeout_ms: int = 2500) -> bool:
    for selector in selectors:
        try:
            locator = page.locator(selector)
            count = locator.count()
            for index in range(min(count, 12)):
                item = locator.nth(index)
                if item.is_visible(timeout=500):
                    item.hover(timeout=timeout_ms)
                    return True
        except Exception:
            continue
    return False


def bounding_box_first_visible(page: Any, selectors: list[str], timeout_ms: int = 1500) -> dict[str, float] | None:
    for selector in selectors:
        try:
            locator = page.locator(selector)
            count = locator.count()
            for index in range(min(count, 12)):
                item = locator.nth(index)
                if item.is_visible(timeout=500):
                    box = item.bounding_box(timeout=timeout_ms)
                    if box:
                        return box
        except Exception:
            continue
    return None


def click_document_more_next_to_share(page: Any) -> None:
    viewport = page.viewport_size or {"width": 1280, "height": 720}
    share_box = bounding_box_first_visible(
        page,
        [
            "button:has-text('分享')",
            "[role='button']:has-text('分享')",
            "text=/分享/",
        ],
    )
    if share_box:
        left = int(share_box["x"])
        right = int(share_box["x"] + share_box["width"])
        top = int(share_box["y"])
        bottom = int(share_box["y"] + share_box["height"])
        print(f"[download] Visually located Share button at x={left}-{right}, y={top}-{bottom}.")
        x = right + 30
        y = int(share_box["y"] + share_box["height"] / 2)
        print(f"[download] Trying DingTalk document More near Share: ({x},{y}) via visual-share.")
        page.mouse.click(x, y)
        return

    x = max(20, viewport["width"] - 59)
    print(f"[download] Share text not found; falling back to top-right More at ({x}, 34).")
    page.mouse.click(x, 34)


def click_export_as_excel_from_bitable_menu(page: Any) -> bool:
    if click_first_visible(
        page,
        [
            "[role='menuitem']:has-text('导出')",
            "[role='option']:has-text('导出')",
            "text=/导出/",
        ],
        timeout_ms=1800,
    ):
        page.wait_for_timeout(500)
        if click_first_visible(
            page,
            [
                "text=/导出为\\s*Excel/i",
                "text=/Excel/i",
            ],
            timeout_ms=1800,
        ):
            return True

    viewport = page.viewport_size or {"width": 1280, "height": 720}
    export_x = max(30, int(viewport["width"] - 229))
    export_y = int(viewport["height"] * 0.593)
    excel_x = max(30, int(viewport["width"] - 489))
    excel_y = export_y
    print(f"[download] Hovering visual Export row at ({export_x},{export_y}).")
    page.mouse.move(export_x, export_y)
    page.wait_for_timeout(650)
    print(f"[download] Clicking visual Export as Excel item at ({excel_x},{excel_y}).")
    page.mouse.click(excel_x, excel_y)
    return True


def click_continue_download_if_needed(page: Any) -> None:
    page.wait_for_timeout(1000)
    if click_first_visible(
        page,
        [
            "button:has-text('继续下载')",
            "[role='button']:has-text('继续下载')",
            "text=/继续下载/",
        ],
        timeout_ms=1500,
    ):
        print("[download] Clicked DingTalk permission continue button by text.")
        page.wait_for_timeout(1200)
        return

    viewport = page.viewport_size or {"width": 1280, "height": 720}
    x = int(viewport["width"] * 0.711)
    y = int(viewport["height"] * 0.554)
    print(f"[download] Clicking visual DingTalk permission continue button at ({x}, {y}).")
    page.mouse.click(x, y)
    page.wait_for_timeout(1200)


def uncheck_attachment_option_if_needed(page: Any) -> None:
    page.wait_for_timeout(800)
    try:
        checkbox = page.locator("label:has-text('包含所选范围中的附件') input[type='checkbox']").first
        if checkbox.count() and checkbox.is_checked(timeout=1000):
            checkbox.click(timeout=1500)
            print("[download] Unchecked attachment option by checkbox selector.")
            return
    except Exception:
        pass

    text_box = bounding_box_first_visible(
        page,
        [
            "text=/包含所选范围中的附件/",
            "label:has-text('包含所选范围中的附件')",
        ],
    )
    if text_box:
        x = int(max(1, text_box["x"] - 18))
        y = int(text_box["y"] + text_box["height"] / 2)
        print(f"[download] Clicking visual attachment checkbox at ({x}, {y}).")
        page.mouse.click(x, y)
        page.wait_for_timeout(500)
        return

    viewport = page.viewport_size or {"width": 1280, "height": 720}
    x = int(viewport["width"] * 0.441)
    y = int(viewport["height"] * 0.393)
    print(f"[download] Attachment checkbox text not found; clicking fallback at ({x}, {y}).")
    page.mouse.click(x, y)
    page.wait_for_timeout(500)


def click_final_export_button(page: Any) -> None:
    if click_first_visible(
        page,
        [
            "button:has-text('导出')",
            "[role='button']:has-text('导出')",
        ],
        timeout_ms=1500,
    ):
        print("[download] Clicked final Export button by text.")
        return
    viewport = page.viewport_size or {"width": 1280, "height": 720}
    x = int(viewport["width"] * 0.768)
    y = int(viewport["height"] * 0.606)
    print(f"[download] Clicking visual final Export button at ({x}, {y}).")
    page.mouse.click(x, y)


def try_bitable_export(page: Any, timeout_ms: int) -> Any | None:
    try:
        print("[download] Trying DingTalk Bitable menu path: Share text -> document More -> Export -> Excel.")
        click_document_more_next_to_share(page)
        page.wait_for_timeout(900)
        click_export_as_excel_from_bitable_menu(page)
        click_continue_download_if_needed(page)
        print("[download] Detected DingTalk Excel export dialog visually.")
        uncheck_attachment_option_if_needed(page)
        with page.expect_download(timeout=timeout_ms) as download_info:
            click_final_export_button(page)
        return download_info.value
    except Exception as exc:
        print(f"[download] DingTalk Bitable export path did not trigger Excel download: {exc}")
        try:
            page.keyboard.press("Escape")
        except Exception:
            pass
        return None


def try_document_more_download(page: Any, timeout_ms: int) -> Any | None:
    """Current DingTalk Sheet layout: top-right more -> 下载为 -> Excel."""
    viewport = page.viewport_size or {"width": 1280, "height": 720}
    menu_x_candidates = [
        max(20, viewport["width"] - 151),
        max(20, viewport["width"] - 145),
        max(20, viewport["width"] - 170),
    ]

    for menu_x in menu_x_candidates:
        try:
            print("[download] Trying DingTalk document menu path: More -> Download as -> Excel.")
            page.mouse.click(menu_x, 34)
            page.wait_for_timeout(800)
            page.mouse.move(max(20, viewport["width"] - 336), 401)
            page.wait_for_timeout(900)
            with page.expect_download(timeout=timeout_ms) as download_info:
                page.mouse.click(max(20, viewport["width"] - 520), 439)
            return download_info.value
        except Exception as exc:
            print(f"[download] Coordinate document menu path did not trigger Excel download: {exc}")
            try:
                page.keyboard.press("Escape")
            except Exception:
                pass

        try:
            page.mouse.click(menu_x, 34)
            page.wait_for_timeout(800)
            opened_download_menu = hover_first_visible(
                page,
                [
                    "text=/下载为/",
                    "[role='menuitem']:has-text('下载为')",
                    "[role='button']:has-text('下载为')",
                ],
                timeout_ms=1800,
            )
            if not opened_download_menu:
                page.keyboard.press("Escape")
                continue
            page.wait_for_timeout(800)
            with page.expect_download(timeout=timeout_ms) as download_info:
                if not click_first_visible(
                    page,
                    [
                        "text=/Excel.*xlsx/i",
                        "text=/Excel/i",
                        "text=/\\.xlsx/i",
                    ],
                    timeout_ms=2500,
                ):
                    raise SyncError("DingTalk download-as menu opened, but Excel item was not found.")
            return download_info.value
        except Exception:
            try:
                page.keyboard.press("Escape")
            except Exception:
                pass
            continue
    return None


def try_auto_click_export(page: Any, timeout_ms: int, export_kind: str = "auto") -> Any | None:
    """Drive the DingTalk export UI.

    DingTalk moves export behind different menus depending on document type and
    rollout version. The strategy is intentionally layered: try direct export
    actions, then open likely menu buttons, then try export actions again.
    """
    page.wait_for_timeout(2500)

    if export_kind == "bitable":
        print("[download] Export kind is bitable; skipping document download probes.")
        return try_bitable_export(page, timeout_ms)

    direct_candidates = [
        "text=/导出为\\s*Excel/i",
        "text=/下载为\\s*Excel/i",
        "text=/导出为Excel/i",
        "text=/下载为Excel/i",
        "text=/Excel/i",
        "[aria-label*='下载']",
        "[title*='下载']",
        "[data-testid*='download']",
    ]

    for selector in direct_candidates:
        try:
            if visible_count(page, selector) == 0:
                continue
            with page.expect_download(timeout=timeout_ms) as download_info:
                page.locator(selector).first.click(timeout=3000)
            return download_info.value
        except Exception:
            continue

    download = try_document_more_download(page, timeout_ms)
    if download is not None:
        return download

    if export_kind in {"auto", "bitable"}:
        download = try_bitable_export(page, timeout_ms)
        if download is not None:
            return download

    menu_triggers = [
        "button:has-text('文件')",
        "[role='button']:has-text('文件')",
        "text=/文件/",
        "button:has-text('更多')",
        "[role='button']:has-text('更多')",
        "[aria-label*='更多']",
        "[title*='更多']",
        "[aria-label*='菜单']",
        "[title*='菜单']",
        "[aria-label*='操作']",
        "[title*='操作']",
        "button:has-text('...')",
        "[role='button']:has-text('...')",
        "button:has-text('⋯')",
        "[role='button']:has-text('⋯')",
        "button:has-text('···')",
        "[role='button']:has-text('···')",
    ]

    for trigger in menu_triggers:
        try:
            locator = page.locator(trigger)
            count = locator.count()
            for index in range(min(count, 8)):
                item = locator.nth(index)
                if not item.is_visible(timeout=500):
                    continue
                item.click(timeout=2500)
                page.wait_for_timeout(800)

                try:
                    with page.expect_download(timeout=timeout_ms) as download_info:
                        if click_export_menu_item(page):
                            pass
                        else:
                            raise SyncError("No export menu item appeared after opening a DingTalk menu.")
                    return download_info.value
                except Exception:
                    # Some menus have a second-level export submenu.
                    if click_first_visible(page, ["text=/导出/i", "text=/下载/i"], timeout_ms=1800):
                        page.wait_for_timeout(800)
                        try:
                            with page.expect_download(timeout=timeout_ms) as download_info:
                                if not click_first_visible(
                                    page,
                                    ["text=/Excel/i", "text=/XLSX/i", "text=/表格/i"],
                                    timeout_ms=1800,
                                ):
                                    raise SyncError("No Excel item appeared in DingTalk export submenu.")
                            return download_info.value
                        except Exception:
                            pass
                page.keyboard.press("Escape")
        except Exception:
            continue

    # Coordinate fallback for current DingTalk document layouts: open top-left
    # menus and then search visible menu text. Kept last because it is visual.
    for x, y in [(28, 24), (56, 24), (96, 24), (1500, 24), (1548, 24)]:
        try:
            page.mouse.click(x, y)
            page.wait_for_timeout(900)
            with page.expect_download(timeout=timeout_ms) as download_info:
                if not click_export_menu_item(page):
                    raise SyncError("No export item after coordinate fallback menu click.")
            return download_info.value
        except Exception:
            try:
                page.keyboard.press("Escape")
            except Exception:
                pass

    return None


def is_dingtalk_login_page(page: Any) -> bool:
    try:
        return "login.dingtalk.com" in page.url or "统一身份认证" in page.title()
    except Exception:
        return False


def try_click_dingtalk_login_authorize(page: Any, login_hint: str) -> bool:
    if not is_dingtalk_login_page(page):
        return False

    if login_hint:
        selectors = [
            f"button:has-text('{login_hint}')",
            f"[role='button']:has-text('{login_hint}')",
            f"text={login_hint}",
        ]
        for selector in selectors:
            try:
                locator = page.locator(selector)
                count = locator.count()
                for index in range(min(count, 5)):
                    item = locator.nth(index)
                    if item.is_visible(timeout=700):
                        item.click(timeout=3000)
                        print(f"[download] Clicked DingTalk login account hint: {login_hint}")
                        return True
            except Exception:
                continue

    try:
        body_text = page.locator("body").inner_text(timeout=1000)
    except Exception:
        body_text = ""
    if "点击头像授权登录" in body_text or "扫码登录" in body_text or "授权登录" in body_text:
        viewport = page.viewport_size or {"width": 1280, "height": 720}
        width = int(viewport["width"])
        height = int(viewport["height"])
        # Current DingTalk auth layout puts the selected account card in the center-right area.
        ratio_points = [
            (0.57, 0.55),
            (0.57, 0.64),
            (0.50, 0.55),
            (0.62, 0.55),
            (0.57, 0.47),
            (0.88, 0.57),
        ]
        points = [(max(20, min(width - 20, int(width * x))), max(80, min(height - 40, int(height * y)))) for x, y in ratio_points]
        for x, y in dict.fromkeys(points):
            try:
                page.mouse.click(x, y)
                page.wait_for_timeout(1200)
                print(
                    "[download] Clicked DingTalk login avatar fallback "
                    f"near account hint: {login_hint or 'default account'} at ({x}, {y})."
                )
                return True
            except Exception:
                continue
    return False


def try_click_dingtalk_doc_login_window(context: Any, login_hint: str) -> bool:
    """Click the DingTalk Docs standalone login popup if it appears."""
    for popup in list(context.pages):
        try:
            if popup.is_closed():
                continue
            title = popup.title()
            body_text = ""
            try:
                body_text = popup.locator("body").inner_text(timeout=1000)
            except Exception:
                pass
            content = f"{title}\n{body_text}"
            if "钉钉文档" not in content or "登录" not in content:
                continue

            popup.bring_to_front()
            for selector in [
                "button:has-text('登录')",
                "[role='button']:has-text('登录')",
                "text=/^登录$/",
            ]:
                try:
                    locator = popup.locator(selector)
                    count = locator.count()
                    for index in range(min(count, 5)):
                        item = locator.nth(index)
                        if item.is_visible(timeout=500):
                            item.click(timeout=3000)
                            print("[download] Clicked DingTalk Docs login popup button.")
                            popup.wait_for_timeout(1000)
                            try_click_dingtalk_login_authorize(popup, login_hint)
                            return True
                except Exception:
                    continue
        except Exception:
            continue
    return False


def process_image_name(pid: int) -> str:
    if os.name != "nt":
        return ""
    try:
        import ctypes
        from ctypes import wintypes

        PROCESS_QUERY_LIMITED_INFORMATION = 0x1000
        handle = ctypes.windll.kernel32.OpenProcess(PROCESS_QUERY_LIMITED_INFORMATION, False, pid)
        if not handle:
            return ""
        try:
            buffer = ctypes.create_unicode_buffer(1024)
            size = wintypes.DWORD(len(buffer))
            ok = ctypes.windll.kernel32.QueryFullProcessImageNameW(handle, 0, buffer, ctypes.byref(size))
            return Path(buffer.value).name.lower() if ok else ""
        finally:
            ctypes.windll.kernel32.CloseHandle(handle)
    except Exception:
        return ""


def click_native_dingtalk_doc_login_popup() -> bool:
    """Click the native-looking DingTalk Docs login popup shown outside the DOM."""
    if os.name != "nt":
        return False

    try:
        import ctypes
        from ctypes import wintypes

        user32 = ctypes.windll.user32
        try:
            user32.SetProcessDPIAware()
        except Exception:
            pass
        target_processes = {"msedge.exe", "msedgewebview2.exe", "chrome.exe", "dingtalk.exe"}
        candidates: list[dict[str, Any]] = []
        foreground = user32.GetForegroundWindow()

        EnumWindowsProc = ctypes.WINFUNCTYPE(wintypes.BOOL, wintypes.HWND, wintypes.LPARAM)

        def callback(hwnd: int, _lparam: int) -> bool:
            if not user32.IsWindowVisible(hwnd):
                return True
            rect = wintypes.RECT()
            if not user32.GetWindowRect(hwnd, ctypes.byref(rect)):
                return True
            width = rect.right - rect.left
            height = rect.bottom - rect.top
            if width < 360 or width > 900 or height < 420 or height > 950:
                return True

            pid = wintypes.DWORD()
            user32.GetWindowThreadProcessId(hwnd, ctypes.byref(pid))
            process_name = process_image_name(int(pid.value))
            if process_name not in target_processes:
                return True

            title_length = user32.GetWindowTextLengthW(hwnd)
            title_buffer = ctypes.create_unicode_buffer(title_length + 1)
            user32.GetWindowTextW(hwnd, title_buffer, title_length + 1)
            title = title_buffer.value
            looks_like_dingtalk_doc_popup = (
                process_name == "dingtalk.exe"
                and not title.strip()
                and 360 <= width <= 720
                and 420 <= height <= 860
            )

            candidates.append(
                {
                    "hwnd": hwnd,
                    "title": title,
                    "left": rect.left,
                    "top": rect.top,
                    "width": width,
                    "height": height,
                    "process_name": process_name,
                    "looks_like_dingtalk_doc_popup": looks_like_dingtalk_doc_popup,
                    "is_foreground": hwnd == foreground,
                }
            )
            return True

        user32.EnumWindows(EnumWindowsProc(callback), 0)
        if not candidates:
            return False
        actionable = [
            candidate
            for candidate in candidates
            if (
                candidate["is_foreground"]
                or candidate["looks_like_dingtalk_doc_popup"]
                or "钉钉" in candidate["title"]
                or "登录" in candidate["title"]
            )
        ]
        if not actionable:
            return False

        def score(candidate: dict[str, Any]) -> tuple[int, int]:
            title = candidate["title"]
            title_score = 2 if "钉钉" in title or "登录" in title else 0
            foreground_score = 3 if candidate["is_foreground"] else 0
            popup_score = 3 if candidate["looks_like_dingtalk_doc_popup"] else 0
            # The popup is usually portrait-ish; prefer that over normal browser windows.
            shape_score = 1 if candidate["height"] >= candidate["width"] else 0
            return (foreground_score + title_score + popup_score + shape_score, candidate["height"] * candidate["width"])

        candidate = sorted(actionable, key=score, reverse=True)[0]
        click_x = int(candidate["left"] + candidate["width"] * 0.5)
        click_y = int(candidate["top"] + candidate["height"] * 0.76)
        hwnd = candidate["hwnd"]
        user32.ShowWindow(hwnd, 9)  # SW_RESTORE
        user32.SetForegroundWindow(hwnd)
        time.sleep(0.2)
        user32.SetCursorPos(click_x, click_y)
        user32.mouse_event(0x0002, 0, 0, 0, 0)  # MOUSEEVENTF_LEFTDOWN
        user32.mouse_event(0x0004, 0, 0, 0, 0)  # MOUSEEVENTF_LEFTUP
        print(
            "[download] Clicked native DingTalk Docs login popup "
            f"at ({click_x}, {click_y}); title={candidate['title']!r}; process={candidate['process_name']}."
        )
        return True
    except Exception as exc:
        print(f"[download] Native DingTalk Docs login popup click failed: {exc}")
        return False


def settle_dingtalk_doc_login_windows(context: Any, login_hint: str, seconds: int = 8) -> None:
    deadline = time.time() + seconds
    while time.time() < deadline:
        clicked = try_click_dingtalk_doc_login_window(context, login_hint)
        if not clicked:
            clicked = click_native_dingtalk_doc_login_popup()
        if not clicked:
            time.sleep(1)
        else:
            time.sleep(2)


def close_stale_dingtalk_preview_pages(context: Any, keep_page: Any) -> None:
    for candidate in list(context.pages):
        try:
            if candidate == keep_page or candidate.is_closed():
                continue
            if "alidocs.dingtalk.com/uni-preview" in candidate.url:
                print(f"[download] Closing stale DingTalk preview tab: {candidate.url}")
                candidate.close()
        except Exception:
            continue


def wait_until_dingtalk_document_page(
    page: Any,
    timeout_sec: int,
    login_hint: str = "",
    auto_login: bool = True,
) -> bool:
    deadline = time.time() + timeout_sec
    last_click_at = 0.0
    while time.time() < deadline:
        if not is_dingtalk_login_page(page):
            return True
        if auto_login and login_hint and time.time() - last_click_at > 5:
            if not try_click_dingtalk_doc_login_window(page.context, login_hint):
                click_native_dingtalk_doc_login_popup()
            if try_click_dingtalk_login_authorize(page, login_hint):
                last_click_at = time.time()
                settle_dingtalk_doc_login_windows(page.context, login_hint, seconds=4)
        page.wait_for_timeout(1000)
    return not is_dingtalk_login_page(page)


def is_dingtalk_desktop_page(page: Any) -> bool:
    try:
        parsed = urlparse(page.url)
        return parsed.netloc.endswith("alidocs.dingtalk.com") and parsed.path.rstrip("/") == "/i/desktop"
    except Exception:
        return False


def is_dingtalk_document_list_page(page: Any, doc_title: str) -> bool:
    if is_dingtalk_desktop_page(page):
        return True
    if not doc_title:
        return False
    try:
        parsed = urlparse(page.url)
        if not parsed.netloc.endswith("alidocs.dingtalk.com"):
            return False
        if parsed.path.startswith("/i/nodes/"):
            return False
        title = page.title()
        body_text = page.locator("body").inner_text(timeout=1000)
        return doc_title in body_text and ("首页" in body_text or "最近" in body_text or title.startswith("文档"))
    except Exception:
        return False


def click_dingtalk_document_title(page: Any, doc_title: str, timeout_ms: int = 15_000) -> Any:
    if not doc_title:
        return page

    page.wait_for_timeout(1500)
    print(f"[download] DingTalk desktop detected; opening document by title: {doc_title}")
    context = page.context
    pages_before = list(context.pages)
    locators = []
    try:
        locators.append(page.get_by_text(doc_title, exact=True))
    except Exception:
        pass
    locators.extend(
        [
            page.locator(f"text={doc_title}"),
            page.locator(f"[title='{doc_title}']"),
        ]
    )

    clicked = False
    last_error: Exception | None = None
    for locator in locators:
        try:
            count = locator.count()
            for index in range(min(count, 5)):
                item = locator.nth(index)
                if item.is_visible(timeout=1000):
                    item.click(timeout=timeout_ms)
                    clicked = True
                    break
            if clicked:
                break
        except Exception as exc:
            last_error = exc

    if not clicked:
        detail = f" Last error: {last_error}" if last_error else ""
        raise SyncError(f"Could not find DingTalk document title on desktop: {doc_title}.{detail}")

    deadline = time.time() + 20
    while time.time() < deadline:
        for candidate in context.pages:
            if candidate not in pages_before and candidate.url != "about:blank":
                candidate.bring_to_front()
                candidate.wait_for_load_state("domcontentloaded", timeout=60_000)
                return candidate
        if not is_dingtalk_desktop_page(page):
            page.wait_for_load_state("domcontentloaded", timeout=60_000)
            return page
        page.wait_for_timeout(500)
    return page


def ensure_dingtalk_target_page(
    page: Any,
    url: str,
    doc_title: str,
    login_hint: str,
    timeout_sec: int,
    auto_login: bool,
) -> Any:
    if auto_login:
        settle_dingtalk_doc_login_windows(page.context, login_hint, seconds=3)
    if is_dingtalk_login_page(page):
        login_mode = "Complete login once" if auto_login else "Complete the login manually"
        print(
            "[download] DingTalk login is required in the opened browser. "
            f"{login_mode}; Excel export will continue automatically."
        )
        if not wait_until_dingtalk_document_page(
            page,
            timeout_sec=timeout_sec,
            login_hint=login_hint,
            auto_login=auto_login,
        ):
            raise SyncError(f"Timed out after {timeout_sec}s waiting for DingTalk login to complete.")
        print("[download] Login completed; reopening configured DingTalk URL.")
        page.goto(url, wait_until="domcontentloaded", timeout=60_000)
        page.wait_for_load_state("domcontentloaded", timeout=60_000)
        page.wait_for_timeout(3000)

    if is_dingtalk_document_list_page(page, doc_title):
        if not doc_title:
            raise SyncError(
                "DingTalk opened the document desktop instead of the target document. "
                "Set DINGTALK_DOC_TITLE to the exact document title to open from the desktop list."
            )
        page = click_dingtalk_document_title(page, doc_title)
        page.wait_for_timeout(4000)

    if auto_login:
        settle_dingtalk_doc_login_windows(page.context, login_hint, seconds=5)
    close_stale_dingtalk_preview_pages(page.context, page)
    print(f"[download] Browser page: {page.title()} | {page.url}")
    return page


def load_browser_cookies(source: str) -> list[dict[str, Any]]:
    try:
        import browser_cookie3
    except ImportError:
        print("[download] browser-cookie3 is not installed; skipping system browser cookie import.")
        return []

    if source not in {"auto", "edge", "chrome"}:
        print(f"[download] Unsupported DINGTALK_BROWSER_COOKIE_SOURCE={source!r}; skipping cookie import.")
        return []

    domains = ["dingtalk.com", "alidocs.dingtalk.com", "login.dingtalk.com"]
    loaders: list[tuple[str, Any]] = []
    if source in {"auto", "edge"}:
        loaders.append(("edge", browser_cookie3.edge))
    if source in {"auto", "chrome"}:
        loaders.append(("chrome", browser_cookie3.chrome))

    cookies_by_key: dict[tuple[str, str, str], dict[str, Any]] = {}
    for loader_name, loader in loaders:
        for domain in domains:
            try:
                jar = loader(domain_name=domain)
            except Exception as exc:
                print(f"[download] Could not read {loader_name} cookies for {domain}: {exc}")
                continue

            for cookie in jar:
                cookie_domain = cookie.domain or domain
                if "dingtalk.com" not in cookie_domain:
                    continue
                item: dict[str, Any] = {
                    "name": cookie.name,
                    "value": cookie.value,
                    "domain": cookie_domain,
                    "path": cookie.path or "/",
                }
                if cookie.expires:
                    item["expires"] = int(cookie.expires)
                cookies_by_key[(item["domain"], item["path"], item["name"])] = item

    cookies = list(cookies_by_key.values())
    if cookies:
        print(f"[download] Imported {len(cookies)} DingTalk cookie(s) from system browser profile.")
    else:
        print("[download] No DingTalk cookies were imported from system browser profiles.")
    return cookies


def download_with_browser(
    url: str,
    doc_title: str,
    login_hint: str,
    output_path: Path,
    profile_dir: Path,
    timeout_sec: int,
    auto_click_export: bool,
    auto_login: bool,
    export_kind: str,
    import_browser_cookies: bool,
    browser_cookie_source: str,
) -> Path:
    try:
        from playwright.sync_api import Error as PlaywrightError
        from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
        from playwright.sync_api import sync_playwright
    except ImportError as exc:
        raise SyncError("Browser download requires Playwright. Run: pip install -r requirements.txt") from exc

    ensure_parent(output_path)
    profile_dir.mkdir(parents=True, exist_ok=True)
    cleanup_automation_browser_processes(profile_dir)
    clear_browser_session_restore(profile_dir)
    mark_browser_profile_clean(profile_dir)
    downloads_dir = output_path.parent.resolve()

    print("[download] Opening browser. Login state is stored locally in the configured profile directory.")
    with sync_playwright() as playwright:
        context = None
        launch_errors: list[str] = []
        for channel in ("msedge", "chrome", None):
            try:
                kwargs: dict[str, Any] = {
                    "headless": False,
                    "accept_downloads": True,
                    "downloads_path": str(downloads_dir),
                    "args": [
                        "--disable-session-crashed-bubble",
                        "--no-first-run",
                        "--no-default-browser-check",
                        "--disable-features=InfiniteSessionRestore",
                    ],
                }
                if channel is not None:
                    kwargs["channel"] = channel
                context = playwright.chromium.launch_persistent_context(str(profile_dir), **kwargs)
                break
            except PlaywrightError as exc:
                launch_errors.append(f"{channel or 'bundled chromium'}: {exc}")

        if context is None:
            details = "\n".join(launch_errors)
            raise SyncError(f"Could not launch a Chromium browser.\n{details}")

        try:
            if import_browser_cookies:
                cookies = load_browser_cookies(browser_cookie_source)
                if cookies:
                    context.add_cookies(cookies)
            restored_pages = list(context.pages)
            page = context.new_page()
            for restored_page in restored_pages:
                try:
                    restored_page.close()
                except Exception:
                    pass
            print(f"[download] Opening configured DingTalk URL: {url}")
            page.goto(url, wait_until="domcontentloaded", timeout=60_000)
            page.bring_to_front()
            page = ensure_dingtalk_target_page(page, url, doc_title, login_hint, timeout_sec, auto_login)

            download = None
            if auto_click_export:
                if auto_login:
                    settle_dingtalk_doc_login_windows(context, login_hint, seconds=5)
                close_stale_dingtalk_preview_pages(context, page)
                print("[download] Trying configured auto-click export selectors...")
                download = try_auto_click_export(page, timeout_ms=15_000, export_kind=export_kind)

            if download is None:
                print(
                    "[download] Automatic export did not find the DingTalk export menu. "
                    "If the browser is waiting on login, finish login; otherwise click Export/Download as Excel."
                )
                try:
                    download = page.wait_for_event("download", timeout=timeout_sec * 1000)
                except PlaywrightTimeoutError as exc:
                    raise SyncError(
                        f"Timed out after {timeout_sec}s waiting for an Excel download from DingTalk."
                    ) from exc

            download.save_as(str(output_path))
            print(f"[download] Saved browser export: {output_path}")
            return output_path
        finally:
            try:
                context.close()
            finally:
                cleanup_automation_browser_processes(profile_dir)


def download_dingtalk_excel(settings: Settings) -> Path:
    last_error: Exception | None = None
    output_path = settings.local_excel_path
    candidate_path = candidate_download_path(output_path)
    for mode in settings.dingtalk_download_mode:
        try:
            if candidate_path.exists():
                candidate_path.unlink()
            if mode == "direct":
                downloaded = try_direct_download(settings.dingtalk_doc_url, candidate_path)
                if downloaded:
                    return commit_downloaded_excel(downloaded, output_path, settings)
                last_error = SyncError("Direct DingTalk download did not return an Excel file.")
            elif mode == "browser":
                downloaded = download_with_browser(
                    settings.dingtalk_doc_url,
                    settings.dingtalk_doc_title,
                    settings.dingtalk_login_hint,
                    candidate_path,
                    settings.dingtalk_browser_profile,
                    settings.dingtalk_browser_timeout_sec,
                    settings.dingtalk_auto_click_export,
                    settings.dingtalk_auto_login,
                    settings.dingtalk_export_kind,
                    settings.dingtalk_import_browser_cookies,
                    settings.dingtalk_browser_cookie_source,
                )
                return commit_downloaded_excel(downloaded, output_path, settings)
        except Exception as exc:
            last_error = exc
            print(f"[download] {mode} mode failed: {exc}")

    if last_error:
        raise SyncError(f"Unable to download DingTalk Excel: {last_error}") from last_error
    raise SyncError("No DingTalk download mode configured.")


def normalize_header(value: Any, index: int, seen: dict[str, int]) -> str:
    name = str(value).strip() if value is not None else ""
    if not name:
        name = f"Column_{index}"
    count = seen.get(name, 0) + 1
    seen[name] = count
    if count > 1:
        name = f"{name}_{count}"
    return name


def excel_value(value: Any) -> Any:
    if isinstance(value, dt.datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, dt.date):
        return dt.datetime.combine(value, dt.time())
    return value


def read_excel_rows(path: Path, sheet_name: str | None, header_row: int) -> tuple[list[str], list[dict[str, Any]]]:
    if not path.exists():
        raise SyncError(f"Excel file does not exist: {path}")
    if header_row < 1:
        raise SyncError("HEADER_ROW must be >= 1.")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise SyncError(f"Worksheet {sheet_name!r} not found. Available: {', '.join(workbook.sheetnames)}")
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook[workbook.sheetnames[0]]

        # DingTalk exports can write a stale <dimension ref="A1"/> even when the sheet
        # contains real data. Recalculate dimensions so streaming reads do not stop at A1.
        if hasattr(worksheet, "reset_dimensions"):
            worksheet.reset_dimensions()

        raw_headers = next(
            worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True),
            None,
        )
        if raw_headers is None:
            raise SyncError(f"No header row found at row {header_row}.")

        seen: dict[str, int] = {}
        headers = [normalize_header(value, idx + 1, seen) for idx, value in enumerate(raw_headers)]
        rows: list[dict[str, Any]] = []

        for raw_row in worksheet.iter_rows(min_row=header_row + 1, max_col=len(headers), values_only=True):
            values = [excel_value(value) for value in raw_row]
            if all(value is None or value == "" for value in values):
                continue
            rows.append(dict(zip(headers, values)))

        print(f"[excel] Loaded {len(rows)} row(s), {len(headers)} column(s) from {path}")
        return headers, rows
    finally:
        workbook.close()


def parse_feishu_url(url: str) -> tuple[str, str]:
    parsed = urlparse(url)
    match = re.search(r"/wiki/([^/?#]+)", parsed.path)
    if not match:
        raise SyncError("FEISHU_BITABLE_URL must be a wiki URL containing /wiki/<token>.")
    wiki_token = unquote(match.group(1))

    query = parse_qs(parsed.query)
    table_id = (query.get("table") or [""])[0]
    if not table_id:
        raise SyncError("FEISHU_BITABLE_URL must include a table=<table_id> query parameter.")
    return wiki_token, table_id


class FeishuClient:
    def __init__(self, app_id: str, app_secret: str, api_base: str) -> None:
        self.api_base = api_base.rstrip("/")
        self.session = requests.Session()
        self.tenant_access_token = self._get_tenant_access_token(app_id, app_secret)

    def _get_tenant_access_token(self, app_id: str, app_secret: str) -> str:
        url = f"{self.api_base}/auth/v3/tenant_access_token/internal"
        response = self.session.post(
            url,
            json={"app_id": app_id, "app_secret": app_secret},
            timeout=30,
        )
        data = self._decode_response(response)
        token = data.get("tenant_access_token")
        if not token:
            raise FeishuAPIError("Feishu token response did not include tenant_access_token.")
        return token

    def request(self, method: str, path: str, **kwargs: Any) -> dict[str, Any]:
        headers = kwargs.pop("headers", {})
        headers["Authorization"] = f"Bearer {self.tenant_access_token}"
        response = self.session.request(
            method,
            f"{self.api_base}{path}",
            headers=headers,
            timeout=60,
            **kwargs,
        )
        return self._decode_response(response)

    @staticmethod
    def _decode_response(response: requests.Response) -> dict[str, Any]:
        try:
            payload = response.json()
        except ValueError as exc:
            raise FeishuAPIError(f"Feishu returned non-JSON response: HTTP {response.status_code}") from exc

        code = payload.get("code", 0)
        if response.status_code >= 400 or code != 0:
            msg = payload.get("msg") or payload.get("message")
            if code == 131006:
                msg = str(msg).rstrip(".。")
                msg = (
                    f"{msg}. The Feishu app can authenticate, but it does not have permission to read "
                    "this wiki/Bitable. Grant the app wiki/docs read permission and Bitable read/write "
                    "permission, reinstall or republish the app, and add the app as a collaborator on "
                    "the target Bitable if required."
                )
            elif code == 131005:
                msg = str(msg).rstrip(".。")
                msg = (
                    f"{msg}. The target wiki/Base node is not visible to this Feishu app. "
                    "If FEISHU_BITABLE_URL is a /wiki/ link, either add the app as a collaborator "
                    "on the target Base/wiki so it can resolve the wiki node, or set "
                    "FEISHU_BITABLE_APP_TOKEN in the corresponding .env file to the real Base token. "
                    "The app still needs Bitable read/write permission on that Base."
                )
            elif code == 91403:
                msg = (
                    f"{msg}. The Bitable app_token/table_id was resolved, but this Feishu app is not "
                    "allowed to access the target Bitable. In Feishu, add the custom app as a collaborator "
                    "or otherwise grant it access to this Base, then confirm the app has Bitable read/write "
                    "scopes and has been republished/reinstalled after permission changes."
                )
            if code:
                raise FeishuAPIError(f"Feishu HTTP {response.status_code} / API error {code}: {msg}")
            raise FeishuAPIError(f"Feishu HTTP {response.status_code}: {json.dumps(payload, ensure_ascii=False)}")
        return payload.get("data", payload)


def resolve_bitable_app_token(client: FeishuClient, wiki_token: str) -> str:
    data = client.request("GET", "/wiki/v2/spaces/get_node", params={"token": wiki_token})
    node = data.get("node") or data
    app_token = node.get("obj_token")
    obj_type = node.get("obj_type")
    if not app_token:
        raise SyncError(f"Could not resolve Bitable app token from wiki node: {json.dumps(data, ensure_ascii=False)}")
    if obj_type and obj_type not in {"bitable", "base"}:
        print(f"[feishu] Warning: wiki node obj_type is {obj_type!r}; continuing with obj_token.")
    return app_token


def list_bitable_fields(client: FeishuClient, app_token: str, table_id: str) -> list[dict[str, Any]]:
    fields: list[dict[str, Any]] = []
    page_token = ""
    while True:
        params: dict[str, Any] = {"page_size": 100}
        if page_token:
            params["page_token"] = page_token
        data = client.request(
            "GET",
            f"/bitable/v1/apps/{app_token}/tables/{table_id}/fields",
            params=params,
        )
        fields.extend(data.get("items", []))
        if not data.get("has_more"):
            break
        page_token = data.get("page_token", "")
    return fields


def create_text_field(client: FeishuClient, app_token: str, table_id: str, field_name: str) -> dict[str, Any]:
    data = client.request(
        "POST",
        f"/bitable/v1/apps/{app_token}/tables/{table_id}/fields",
        json={"field_name": field_name, "type": 1, "property": {}},
    )
    return data.get("field") or data


def list_record_ids(client: FeishuClient, app_token: str, table_id: str) -> list[str]:
    record_ids: list[str] = []
    page_token = ""
    while True:
        params: dict[str, Any] = {"page_size": 500}
        if page_token:
            params["page_token"] = page_token
        data = client.request(
            "GET",
            f"/bitable/v1/apps/{app_token}/tables/{table_id}/records",
            params=params,
        )
        for item in data.get("items", []):
            record_id = item.get("record_id")
            if record_id:
                record_ids.append(record_id)
        if not data.get("has_more"):
            break
        page_token = data.get("page_token", "")
    return record_ids


def chunked(items: list[Any], size: int) -> list[list[Any]]:
    return [items[index : index + size] for index in range(0, len(items), size)]


def delete_records(client: FeishuClient, app_token: str, table_id: str, record_ids: list[str]) -> None:
    for chunk in chunked(record_ids, 500):
        client.request(
            "POST",
            f"/bitable/v1/apps/{app_token}/tables/{table_id}/records/batch_delete",
            json={"records": chunk},
        )
        print(f"[feishu] Deleted {len(chunk)} existing record(s).")


def parse_datetime_text(value: str) -> dt.datetime | None:
    text = value.strip()
    if not text:
        return None

    normalized = text.replace("/", "-").replace(".", "-")
    formats = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%Y-%m",
        "%Y%m%d",
    ]
    for fmt in formats:
        try:
            return dt.datetime.strptime(normalized, fmt)
        except ValueError:
            continue

    try:
        return dt.datetime.fromisoformat(text)
    except ValueError:
        return None


def to_feishu_date_ms(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, dt.datetime):
        parsed = value
    elif isinstance(value, dt.date):
        parsed = dt.datetime.combine(value, dt.time())
    elif isinstance(value, (int, float)):
        # If a timestamp is already in milliseconds, keep it. Otherwise treat it as seconds.
        return int(value if value > 10_000_000_000 else value * 1000)
    else:
        parsed = parse_datetime_text(str(value))
        if parsed is None:
            return None
    # Feishu date fields are returned as the local midnight timestamp. For a
    # China tenant that means 2025-10-13 is 2025-10-12T16:00:00Z.
    return int(parsed.timestamp() * 1000)


def split_multi_value(value: Any) -> list[str]:
    if value is None or value == "":
        return []
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    return [part.strip() for part in re.split(r"[,，;；\n]+", str(value)) if part.strip()]


def normalize_building(value: Any) -> Any:
    if value is None or value == "":
        return value
    text = str(value).strip()
    aliases = {
        "园区": "园区级",
        "园区级": "园区级",
        "110": "110站",
        "110站": "110站",
    }
    if text in aliases:
        return aliases[text]
    if re.fullmatch(r"[A-Za-z]", text):
        return f"{text.upper()}楼"
    return text


def apply_transform(value: Any, transform: str | None) -> Any:
    if not transform:
        return value
    if transform == "building":
        return normalize_building(value)
    raise SyncError(f"Unsupported field mapping transform: {transform}")


def collect_template_values(template: str) -> set[str]:
    names: set[str] = set()
    for expression in re.findall(r"\{([^{}]+)\}", template):
        name = expression.split(":", 1)[0]
        if name:
            names.add(name)
    return names


def render_template_value(template: str, values: dict[str, Any]) -> str:
    def replace(match: re.Match[str]) -> str:
        expression = match.group(1)
        parts = expression.split(":")
        name = parts[0]
        transform: str | None = None
        fmt: str | None = None
        for part in parts[1:]:
            if part in {"building"}:
                transform = part
            else:
                fmt = part
        value = apply_transform(values.get(name), transform)
        return format_key_value(value, fmt)

    return re.sub(r"\{([^{}]+)\}", replace, template)


def to_bool(value: Any) -> bool | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    text = str(value).strip().lower()
    if text in {"true", "yes", "y", "1", "是", "对", "勾选"}:
        return True
    if text in {"false", "no", "n", "0", "否", "不", "未勾选"}:
        return False
    return None


def to_number(value: Any) -> int | float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip().replace(",", "")
    try:
        number = float(text)
    except ValueError:
        return None
    return int(number) if number.is_integer() else number


def convert_field_value(value: Any, field: dict[str, Any]) -> Any:
    if value is None or value == "":
        return None

    field_type = field.get("type")
    if field_type == 2:  # number
        return to_number(value)
    if field_type == 3:  # single select
        return str(value).strip()
    if field_type == 4:  # multi-select
        return split_multi_value(value)
    if field_type == 5:  # date
        return to_feishu_date_ms(value)
    if field_type == 7:  # checkbox
        return to_bool(value)
    if field_type == 11:  # user
        if isinstance(value, list):
            return value or None
        if isinstance(value, dict):
            return [value]
    if field_type == 15:  # hyperlink
        text = str(value).strip()
        if text.startswith(("http://", "https://")):
            return {"text": text, "link": text}
        return {"text": text, "link": text}
    if isinstance(value, dt.datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, dt.date):
        return value.strftime("%Y-%m-%d")
    return value


def load_field_mapping(path: Path | None) -> dict[str, Any]:
    if path is None:
        return {}
    if not path.exists():
        raise SyncError(f"FIELD_MAPPING_FILE does not exist: {path}")
    with path.open("r", encoding="utf-8") as file:
        data = json.load(file)
    if not isinstance(data, dict):
        raise SyncError("FIELD_MAPPING_FILE must contain a JSON object.")
    return data


def source_headers_from_mapping(field_mapping: dict[str, Any]) -> set[str]:
    headers: set[str] = set(mapping_entries(field_mapping).keys())
    key_config = field_mapping.get("key") or {}
    if isinstance(key_config, dict) and key_config.get("template"):
        headers.update(collect_template_values(str(key_config["template"])))
    for target_config in (field_mapping.get("derived_fields") or {}).values():
        if isinstance(target_config, dict):
            source = target_config.get("source")
            if source:
                headers.add(str(source))
            template = target_config.get("template")
            if template:
                headers.update(collect_template_values(str(template)))
    return headers


def mapping_entries(field_mapping: dict[str, Any]) -> dict[str, dict[str, str | None]]:
    raw = field_mapping.get("mappings", {})
    if not raw:
        return {}
    if not isinstance(raw, dict):
        raise SyncError("field mapping 'mappings' must be an object.")

    entries: dict[str, dict[str, str | None]] = {}
    for source, target_config in raw.items():
        if isinstance(target_config, str):
            entries[source] = {"target": target_config, "transform": None}
        elif isinstance(target_config, dict) and target_config.get("target"):
            entries[source] = {
                "target": str(target_config["target"]),
                "transform": target_config.get("transform"),
            }
        else:
            raise SyncError(f"Invalid field mapping for source column {source!r}.")
    return entries


def validate_downloaded_excel(path: Path, settings: Settings) -> None:
    try:
        headers, rows = read_excel_rows(path, settings.sheet_name, settings.header_row)
    except Exception as exc:
        raise SyncError(
            "Downloaded Excel does not match the configured DingTalk source. "
            f"The candidate file was kept at {path}; the existing local Excel was not replaced. "
            f"Reason: {exc}"
        ) from exc

    field_mapping = load_field_mapping(settings.field_mapping_file)
    expected_source_headers = list(source_headers_from_mapping(field_mapping))
    missing_headers = [header for header in expected_source_headers if header not in headers]
    if missing_headers:
        raise SyncError(
            "Downloaded Excel is not the expected source workbook. "
            f"Missing required column(s): {', '.join(missing_headers)}. "
            f"The candidate file was kept at {path}; the existing local Excel was not replaced."
        )
    if not rows:
        raise SyncError(
            "Downloaded Excel has no data rows. "
            f"The candidate file was kept at {path}; the existing local Excel was not replaced."
        )
    print(
        f"[download] Verified Excel source: sheet={settings.sheet_name or '<active>'}, "
        f"rows={len(rows)}, columns={len(headers)}."
    )


def commit_downloaded_excel(downloaded_path: Path, output_path: Path, settings: Settings) -> Path:
    validate_downloaded_excel(downloaded_path, settings)
    ensure_parent(output_path)
    last_error: Exception | None = None
    for attempt in range(1, 21):
        try:
            downloaded_path.replace(output_path)
            break
        except PermissionError as exc:
            last_error = exc
            print(f"[download] Waiting for Excel candidate file lock to release ({attempt}/20)...")
            time.sleep(1)
    else:
        try:
            shutil.copy2(downloaded_path, output_path)
            print(f"[download] Copied locked Excel candidate into local Excel: {output_path}")
            return output_path
        except Exception:
            print(
                f"[download] Could not replace {output_path} because Windows is holding a file lock. "
                f"Continuing with validated candidate: {downloaded_path}. Last error: {last_error}"
            )
            return downloaded_path
    print(f"[download] Replaced local Excel after validation: {output_path}")
    return output_path


def field_display_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        parts: list[str] = []
        for item in value:
            if isinstance(item, dict):
                text = item.get("text") or item.get("name") or item.get("value")
                if text is not None:
                    parts.append(str(text))
            elif item is not None:
                parts.append(str(item))
        return ",".join(parts)
    if isinstance(value, dict):
        return str(value.get("text") or value.get("name") or value.get("value") or "")
    return str(value)


def format_key_value(value: Any, fmt: str | None = None) -> str:
    if fmt and isinstance(value, str):
        parsed_text = parse_datetime_text(value)
        if parsed_text is not None:
            python_fmt = fmt.replace("YYYY", "%Y").replace("MM", "%m").replace("DD", "%d")
            return parsed_text.strftime(python_fmt)
    if isinstance(value, (dt.datetime, dt.date)):
        if fmt:
            python_fmt = fmt.replace("YYYY", "%Y").replace("MM", "%m").replace("DD", "%d")
            return value.strftime(python_fmt)
        return value.strftime("%Y/%m/%d")
    if isinstance(value, (int, float)) and fmt:
        date_ms = to_feishu_date_ms(value)
        if date_ms is not None:
            parsed = dt.datetime.fromtimestamp(date_ms / 1000, tz=dt.timezone.utc)
            python_fmt = fmt.replace("YYYY", "%Y").replace("MM", "%m").replace("DD", "%d")
            return parsed.strftime(python_fmt)
    return field_display_value(value)


def render_record_key(template: str, values: dict[str, Any]) -> str:
    def replace(match: re.Match[str]) -> str:
        expression = match.group(1)
        if ":" in expression:
            name, fmt = expression.split(":", 1)
        else:
            name, fmt = expression, None
        return format_key_value(values.get(name), fmt)

    return re.sub(r"\{([^{}]+)\}", replace, template)


def split_person_names(value: Any) -> list[str]:
    text = field_display_value(value).strip()
    if not text:
        return []
    return [item.strip() for item in re.split(r"[,，、;/；\s]+", text) if item.strip()]


def feishu_user_value(value: Any, field_mapping: dict[str, Any]) -> list[dict[str, str]]:
    lookup = field_mapping.get("_feishu_user_lookup") or {}
    users: list[dict[str, str]] = []
    for name in split_person_names(value):
        user_id = lookup.get(name)
        if user_id:
            users.append({"id": str(user_id)})
    return users


def apply_mapping_transform(value: Any, transform: str | None, field_mapping: dict[str, Any]) -> Any:
    if transform == "feishu_user":
        return feishu_user_value(value, field_mapping)
    return apply_transform(value, transform)


def prepare_row_fields(
    row: dict[str, Any],
    headers: list[str],
    fields_by_name: dict[str, dict[str, Any]],
    field_mapping: dict[str, Any],
) -> tuple[dict[str, Any], str | None]:
    entries = mapping_entries(field_mapping)
    defaults = field_mapping.get("defaults", {}) or {}
    if not isinstance(defaults, dict):
        raise SyncError("field mapping 'defaults' must be an object.")

    raw_fields: dict[str, Any] = {}
    field_transforms = field_mapping.get("field_transforms", {}) or {}
    if not isinstance(field_transforms, dict):
        raise SyncError("field mapping 'field_transforms' must be an object.")

    if entries:
        for source, config in entries.items():
            target = config["target"]
            if not target:
                continue
            raw_fields[target] = apply_mapping_transform(row.get(source), config.get("transform"), field_mapping)
    else:
        for header in headers:
            raw_fields[header] = apply_mapping_transform(row.get(header), field_transforms.get(header), field_mapping)

    for target, value in defaults.items():
        raw_fields.setdefault(target, value)

    derived_fields = field_mapping.get("derived_fields", {}) or {}
    if not isinstance(derived_fields, dict):
        raise SyncError("field mapping 'derived_fields' must be an object.")
    for target, config in derived_fields.items():
        if isinstance(config, str):
            raw_fields[str(target)] = render_template_value(config, {**row, **raw_fields})
        elif isinstance(config, dict):
            if config.get("template"):
                raw_fields[str(target)] = render_template_value(str(config["template"]), {**row, **raw_fields})
            elif config.get("source"):
                raw_fields[str(target)] = apply_mapping_transform(
                    row.get(str(config["source"])),
                    config.get("transform"),
                    field_mapping,
                )
        else:
            raise SyncError(f"Invalid derived field mapping for {target!r}.")

    converted_fields: dict[str, Any] = {}
    for target, value in raw_fields.items():
        field = fields_by_name.get(target)
        if not field or field.get("type") in READ_ONLY_FIELD_TYPES:
            continue
        converted = convert_field_value(value, field)
        if converted is not None:
            converted_fields[target] = converted

    key_config = field_mapping.get("key") or {}
    record_key = None
    if isinstance(key_config, dict) and key_config.get("template"):
        record_key = render_record_key(str(key_config["template"]), raw_fields)
    return converted_fields, record_key


def prepare_records(
    headers: list[str],
    rows: list[dict[str, Any]],
    fields_by_name: dict[str, dict[str, Any]],
    field_mapping: dict[str, Any] | None = None,
) -> list[dict[str, dict[str, Any]]]:
    field_mapping = field_mapping or {}
    records: list[dict[str, dict[str, Any]]] = []
    for row in rows:
        fields, _ = prepare_row_fields(row, headers, fields_by_name, field_mapping)
        if fields:
            records.append({"fields": fields})
    if not records:
        raise SyncError("No writable Bitable fields match the Excel rows and field mapping.")
    return records


def create_records(client: FeishuClient, app_token: str, table_id: str, records: list[dict[str, Any]]) -> int:
    created = 0
    for chunk in chunked(records, 500):
        client.request(
            "POST",
            f"/bitable/v1/apps/{app_token}/tables/{table_id}/records/batch_create",
            json={"records": chunk},
        )
        created += len(chunk)
        print(f"[feishu] Created {len(chunk)} record(s).")
    return created


def update_records(client: FeishuClient, app_token: str, table_id: str, records: list[dict[str, Any]]) -> int:
    updated = 0
    for chunk in chunked(records, 500):
        client.request(
            "POST",
            f"/bitable/v1/apps/{app_token}/tables/{table_id}/records/batch_update",
            json={"records": chunk},
        )
        updated += len(chunk)
        print(f"[feishu] Updated {len(chunk)} record(s).")
    return updated


def list_feishu_chat_members(client: FeishuClient, chat_id: str) -> list[dict[str, Any]]:
    members: list[dict[str, Any]] = []
    page_token = ""
    while True:
        params: dict[str, Any] = {"member_id_type": "open_id", "page_size": 100}
        if page_token:
            params["page_token"] = page_token
        data = client.request("GET", f"/im/v1/chats/{chat_id}/members", params=params)
        members.extend(data.get("items", []))
        if not data.get("has_more"):
            break
        page_token = data.get("page_token", "")
    return members


def build_feishu_user_lookup(client: FeishuClient, chat_id: str | None) -> dict[str, str]:
    if not chat_id:
        return {}
    try:
        members = list_feishu_chat_members(client, chat_id)
    except Exception as exc:
        print(f"[feishu] Could not load Feishu member lookup chat; user fields will be left blank. Reason: {exc}")
        return {}
    lookup: dict[str, str] = {}
    for member in members:
        member_id = member.get("member_id") or member.get("open_id") or member.get("user_id")
        if not member_id:
            continue
        for key in ("name", "en_name", "nickname"):
            name = str(member.get(key) or "").strip()
            if name:
                lookup[name] = str(member_id)
    print(f"[feishu] Loaded {len(lookup)} Feishu member name mapping(s).")
    return lookup


def extract_record_key(record: dict[str, Any], key_field: str) -> str:
    return field_display_value((record.get("fields") or {}).get(key_field)).strip()


def is_empty_feishu_value(value: Any) -> bool:
    if value is None or value == "":
        return True
    if isinstance(value, list):
        return len(value) == 0
    if isinstance(value, dict):
        return not field_display_value(value).strip()
    return False


def list_records_for_upsert(
    client: FeishuClient,
    app_token: str,
    table_id: str,
    key_field: str,
) -> dict[str, str]:
    records_by_key: dict[str, str] = {}
    page_token = ""
    while True:
        params: dict[str, Any] = {"page_size": 500}
        if page_token:
            params["page_token"] = page_token
        data = client.request(
            "GET",
            f"/bitable/v1/apps/{app_token}/tables/{table_id}/records",
            params=params,
        )
        for item in data.get("items", []):
            record_id = item.get("record_id")
            record_key = extract_record_key(item, key_field)
            if record_id and record_key:
                records_by_key[record_key] = record_id
        if not data.get("has_more"):
            break
        page_token = data.get("page_token", "")
    return records_by_key


def list_records_for_keyed_sync(
    client: FeishuClient,
    app_token: str,
    table_id: str,
    key_field: str,
) -> dict[str, dict[str, Any]]:
    records_by_key: dict[str, dict[str, Any]] = {}
    page_token = ""
    while True:
        params: dict[str, Any] = {"page_size": 500}
        if page_token:
            params["page_token"] = page_token
        data = client.request(
            "GET",
            f"/bitable/v1/apps/{app_token}/tables/{table_id}/records",
            params=params,
        )
        for item in data.get("items", []):
            record_id = item.get("record_id")
            record_key = extract_record_key(item, key_field)
            if record_id and record_key:
                records_by_key[record_key] = item
        if not data.get("has_more"):
            break
        page_token = data.get("page_token", "")
    return records_by_key


def upsert_records(
    client: FeishuClient,
    app_token: str,
    table_id: str,
    headers: list[str],
    rows: list[dict[str, Any]],
    fields_by_name: dict[str, dict[str, Any]],
    field_mapping: dict[str, Any],
    update_existing: bool = True,
) -> tuple[int, int, int]:
    key_config = field_mapping.get("key") or {}
    key_field = key_config.get("target_field") if isinstance(key_config, dict) else None
    if not key_field:
        raise SyncError("Keyed sync requires field_mapping.json key.target_field.")

    existing = list_records_for_keyed_sync(client, app_token, table_id, str(key_field))
    fill_empty_targets = field_mapping.get("fill_empty_existing_targets", []) or []
    if not isinstance(fill_empty_targets, list):
        raise SyncError("field mapping 'fill_empty_existing_targets' must be a list.")
    fill_empty_targets = [str(target) for target in fill_empty_targets]
    update_existing_targets = field_mapping.get("update_existing_targets", []) or []
    if not isinstance(update_existing_targets, list):
        raise SyncError("field mapping 'update_existing_targets' must be a list.")
    update_existing_targets = [str(target) for target in update_existing_targets]
    to_update: list[dict[str, Any]] = []
    to_create: list[dict[str, Any]] = []
    seen_keys: set[str] = set()
    skipped = 0

    for row in rows:
        fields, record_key = prepare_row_fields(row, headers, fields_by_name, field_mapping)
        if not fields:
            skipped += 1
            continue
        if not record_key:
            raise SyncError("Could not render record key for one of the Excel rows.")
        if record_key in seen_keys:
            print(f"[feishu] Skipping duplicate source key: {record_key}")
            skipped += 1
            continue
        seen_keys.add(record_key)
        existing_record = existing.get(record_key)
        if existing_record:
            record_id = existing_record.get("record_id")
            if update_existing:
                to_update.append({"record_id": record_id, "fields": fields})
            else:
                existing_fields = existing_record.get("fields") or {}
                fill_fields = {
                    target: fields[target]
                    for target in fill_empty_targets
                    if target in fields and is_empty_feishu_value(existing_fields.get(target))
                }
                update_fields = {
                    target: fields[target]
                    for target in update_existing_targets
                    if target in fields
                }
                update_fields.update(fill_fields)
                if update_fields and record_id:
                    to_update.append({"record_id": record_id, "fields": update_fields})
                else:
                    skipped += 1
        else:
            to_create.append({"fields": fields})

    updated = update_records(client, app_token, table_id, to_update) if to_update else 0
    created = create_records(client, app_token, table_id, to_create) if to_create else 0
    return updated, created, skipped


def sync_to_feishu(settings: Settings, headers: list[str], rows: list[dict[str, Any]]) -> SyncResult:
    if not rows:
        raise SyncError("Excel has no data rows. Refusing to clear or write the Feishu table.")

    wiki_token, table_id = parse_feishu_url(settings.feishu_bitable_url)
    client = FeishuClient(settings.feishu_app_id, settings.feishu_app_secret, settings.feishu_api_base)
    if settings.feishu_bitable_app_token:
        app_token = settings.feishu_bitable_app_token
        print(f"[feishu] Using FEISHU_BITABLE_APP_TOKEN. Table: {table_id}")
    else:
        app_token = resolve_bitable_app_token(client, wiki_token)
        print(f"[feishu] Resolved wiki token to Bitable app token. Table: {table_id}")

    fields = list_bitable_fields(client, app_token, table_id)
    fields_by_name = {field.get("field_name"): field for field in fields if field.get("field_name")}

    field_mapping = load_field_mapping(settings.field_mapping_file)
    if "feishu_user" in json.dumps(field_mapping, ensure_ascii=False):
        field_mapping = dict(field_mapping)
        field_mapping["_feishu_user_lookup"] = build_feishu_user_lookup(
            client,
            settings.feishu_member_lookup_chat_id,
        )
    mapped_targets = {
        config["target"]
        for config in mapping_entries(field_mapping).values()
        if config.get("target")
    }
    mapped_targets.update((field_mapping.get("defaults") or {}).keys())
    mapped_targets.update((field_mapping.get("derived_fields") or {}).keys())
    mapped_targets.update(field_mapping.get("update_existing_targets") or [])
    mapped_targets.update(field_mapping.get("fill_empty_existing_targets") or [])
    headers_to_check = list(mapped_targets) if mapped_targets else headers

    missing_headers = [header for header in headers_to_check if header not in fields_by_name]
    if missing_headers and settings.create_missing_fields:
        for header in missing_headers:
            created = create_text_field(client, app_token, table_id, header)
            fields_by_name[created.get("field_name", header)] = created or {"field_name": header, "type": 1}
            print(f"[feishu] Created missing text field: {header}")
    elif missing_headers:
        print(f"[feishu] Skipping fields missing from Bitable: {', '.join(missing_headers)}")

    if settings.sync_mode in {"insert_missing", "upsert"}:
        update_existing = settings.sync_mode == "upsert"
        updated, created, skipped = upsert_records(
            client,
            app_token,
            table_id,
            headers,
            rows,
            fields_by_name,
            field_mapping,
            update_existing=update_existing,
        )
        if update_existing:
            print(
                f"[done] Upserted records to Feishu Bitable. "
                f"Updated: {updated}, created: {created}, skipped: {skipped}."
            )
        else:
            print(
                f"[done] Insert-missing sync finished. "
                f"Created: {created}, filled empty existing fields: {updated}, "
                f"skipped existing/duplicate rows: {skipped}."
            )
        return SyncResult(
            title=settings.dingtalk_doc_title or "钉钉文档同步",
            excel_rows=len(rows),
            created=created,
            updated=updated,
            skipped=skipped,
            mode=settings.sync_mode,
            target_url=settings.feishu_bitable_url,
        )

    records = prepare_records(headers, rows, fields_by_name, field_mapping)

    if settings.sync_mode == "replace":
        record_ids = list_record_ids(client, app_token, table_id)
        if record_ids:
            delete_records(client, app_token, table_id, record_ids)
        else:
            print("[feishu] Target table has no existing records.")

    created = create_records(client, app_token, table_id, records)
    print(f"[done] Synced {created} record(s) to Feishu Bitable.")
    return SyncResult(
        title=settings.dingtalk_doc_title or "钉钉文档同步",
        excel_rows=len(rows),
        created=created,
        updated=0,
        skipped=0,
        mode=settings.sync_mode,
        target_url=settings.feishu_bitable_url,
    )


def print_dry_run(headers: list[str], rows: list[dict[str, Any]]) -> None:
    print("[dry-run] Excel headers:")
    print(json.dumps(headers, ensure_ascii=False, indent=2, default=str))
    print("[dry-run] First rows:")
    print(json.dumps(rows[:3], ensure_ascii=False, indent=2, default=str))
    print(f"[dry-run] Parsed {len(rows)} row(s). No Feishu API calls were made.")


def build_success_notification(settings: Settings, result: SyncResult) -> str:
    lines = [
        f"{result.title} 同步成功",
        f"成功录入 {result.created} 条问题记录。",
        f"Excel 行数：{result.excel_rows}；跳过已有/重复：{result.skipped}；更新已有字段：{result.updated}。",
    ]
    if result.target_url:
        lines.append(f"飞书多维表：{result.target_url}")
    return "\n".join(lines)


def send_feishu_webhook_message(webhook_url: str, text: str) -> None:
    response = requests.post(webhook_url, json={"msg_type": "text", "content": {"text": text}}, timeout=30)
    if response.status_code >= 400:
        raise SyncError(f"Feishu webhook returned HTTP {response.status_code}: {response.text[:300]}")
    payload = response.json()
    if payload.get("code", 0) not in {0, None}:
        raise SyncError(f"Feishu webhook error: {json.dumps(payload, ensure_ascii=False)}")


def send_feishu_im_message(settings: Settings, text: str) -> None:
    app_id = settings.feishu_notify_app_id or settings.feishu_app_id
    app_secret = settings.feishu_notify_app_secret or settings.feishu_app_secret
    receive_id = settings.feishu_notify_chat_id or settings.feishu_notify_open_id
    if not receive_id:
        raise SyncError("Missing Feishu notification receive id.")
    receive_id_type = "chat_id" if settings.feishu_notify_chat_id else "open_id"
    client = FeishuClient(app_id, app_secret, settings.feishu_api_base)
    client.request(
        "POST",
        "/im/v1/messages",
        params={"receive_id_type": receive_id_type},
        json={
            "receive_id": receive_id,
            "msg_type": "text",
            "content": json.dumps({"text": text}, ensure_ascii=False),
        },
    )


def notify_sync_success(settings: Settings, result: SyncResult) -> None:
    if not settings.feishu_notify_enabled:
        return
    if not (settings.feishu_notify_webhook_url or settings.feishu_notify_chat_id or settings.feishu_notify_open_id):
        print("[notify] Skipped Feishu notification; configure FEISHU_NOTIFY_WEBHOOK_URL or FEISHU_NOTIFY_CHAT_ID.")
        return
    text = build_success_notification(settings, result)
    try:
        if settings.feishu_notify_webhook_url:
            send_feishu_webhook_message(settings.feishu_notify_webhook_url, text)
            print("[notify] Sent Feishu webhook notification.")
        else:
            send_feishu_im_message(settings, text)
            print("[notify] Sent Feishu IM notification.")
    except Exception as exc:
        print(f"[notify] Failed to send Feishu notification: {exc}")


def main() -> int:
    args = parse_args()
    try:
        settings = load_settings(args)

        excel_path = settings.local_excel_path
        if args.skip_download:
            print(f"[download] Skipped. Using local Excel: {excel_path}")
        else:
            excel_path = download_dingtalk_excel(settings)

        if args.download_only:
            print("[done] Download-only mode finished.")
            return 0

        headers, rows = read_excel_rows(excel_path, settings.sheet_name, settings.header_row)
        if args.dry_run:
            print_dry_run(headers, rows)
            return 0

        result = sync_to_feishu(settings, headers, rows)
        notify_sync_success(settings, result)
        return 0
    except KeyboardInterrupt:
        print("\n[stopped] Interrupted by user.")
        return 130
    except Exception as exc:
        print(f"[error] {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
